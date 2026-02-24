"""
Copyright 2026 Huawei Technologies Co., Ltd

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

import os
import traceback
import uuid
from typing import List, Optional, Union, Any, Tuple

from backend.services.llm_service import LLMService
from backend.services.document_service import DocumentService
from backend.models.question import Question
from backend.models.document import Document
from backend.models.constants import (
    DOCUMENT_STATUS_QUESTIONS_GENERATED,
    QUESTION_STATUS_UNANSWERED,
    QUESTION_STATUS_ANSWERED,
    DEFAULT_QUESTIONS_PER_CHUNK,
    DEFAULT_QUESTIONS_DIR,
    DEFAULT_GENERATE_PROMPT,
)
from backend.utils.file_utils import save_file, read_file
from backend.utils.logger import init_logger

logger = init_logger(__name__)

# ==================== XLSX导入相关常量 ====================
HEADER_KEYWORDS = {
    '问题', 'question', '内容', 'content',
    '答案', 'answer', '思考过程', 'chain', 'thought'
}
CONTENT_PREVIEW_LENGTH = 50


class QuestionService:
    """
    问题服务类，负责问题的生成、管理与答案生成

    Attributes:
        project_dir: 项目根目录路径
        llm_service: 语言模型服务实例
        questions_dir: 问题存储目录
        system_prompt: 系统提示词
    """

    def __init__(
        self,
        project_dir: str,
        llm_service: LLMService,
        system_prompt: str = DEFAULT_GENERATE_PROMPT,
    ):
        """
        初始化问题管理器

        Args:
            project_dir: 项目根目录路径
            llm_service: 语言模型服务实例
            system_prompt: 系统提示词
        """
        self.project_dir = project_dir
        self.llm_service = llm_service
        self.questions_dir = os.path.join(project_dir, DEFAULT_QUESTIONS_DIR)
        os.makedirs(self.questions_dir, exist_ok=True)

        # 系统提示词，用于控制LLM生成问题的风格
        self.system_prompt = (
            system_prompt
            if system_prompt != DEFAULT_GENERATE_PROMPT
            else llm_service.system_prompt
        )

    @staticmethod
    def _validate_xlsx_dependencies() -> bool:
        """
        验证XLSX处理依赖是否可用

        Returns:
            bool: 依赖是否可用
        """
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            logger.error("错误: 缺少必需的包 'openpyxl'，请运行: pip install openpyxl")
            return False

    @staticmethod
    def _validate_xlsx_file(xlsx_file_path: str) -> bool:
        """
        验证XLSX文件是否存在

        Args:
            xlsx_file_path: 文件路径

        Returns:
            bool: 文件是否有效
        """
        if not os.path.exists(xlsx_file_path):
            logger.error(f"错误: XLSX文件不存在 - {xlsx_file_path}")
            return False
        return True

    @staticmethod
    def _load_xlsx_workbook(xlsx_file_path: str) -> Optional[Any]:
        """
        加载XLSX工作簿

        Args:
            xlsx_file_path: 文件路径

        Returns:
            Optional[Any]: 工作簿对象或None
        """
        try:
            import openpyxl
            from openpyxl.utils.exceptions import InvalidFileException

            try:
                return openpyxl.load_workbook(xlsx_file_path)
            except InvalidFileException:
                logger.error(f"错误: 无效的Excel文件格式 - {xlsx_file_path}")
                return None
        except Exception as e:
            logger.error(f"错误: 加载Excel文件失败 - {e}")
            return None

    @staticmethod
    def _detect_data_start_row(worksheet: Any) -> int:
        """
        检测数据起始行（跳过标题行）

        Args:
            worksheet: 工作表对象

        Returns:
            int: 数据起始行号
        """
        if worksheet.max_row < 1:
            return 1

        first_row_values = [cell.value for cell in worksheet[1]]
        has_header = any(
            val and str(val).lower().strip() in HEADER_KEYWORDS
            for val in first_row_values
            if val is not None
        )

        return 2 if has_header else 1

    @staticmethod
    def _extract_row_data(row: Tuple) -> dict:
        """
        从行数据中提取问题、答案和思维链

        Args:
            row: 行数据元组

        Returns:
            dict: 包含content、answer、chain_of_thought的字典
        """
        content = str(row[0]).strip() if row[0] is not None else ""
        answer = (
            str(row[1]).strip()
            if len(row) > 1 and row[1] is not None
            else None
        )
        chain_of_thought = (
            str(row[2]).strip()
            if len(row) > 2 and row[2] is not None
            else None
        )

        return {
            'content': content,
            'answer': answer,
            'chain_of_thought': chain_of_thought,
        }

    @staticmethod
    def _determine_question_status(row_data: dict) -> int:
        """
        根据数据确定问题状态

        Args:
            row_data: 行数据字典

        Returns:
            int: 问题状态码
        """
        if row_data['chain_of_thought'] and row_data['answer']:
            return QUESTION_STATUS_ANSWERED
        elif row_data['answer']:
            return QUESTION_STATUS_ANSWERED
        return QUESTION_STATUS_UNANSWERED

    @staticmethod
    def _print_import_progress(count: int, content: str) -> None:
        """
        打印导入进度

        Args:
            count: 已处理数量
            content: 问题内容
        """
        preview = content[:CONTENT_PREVIEW_LENGTH]
        suffix = '...' if len(content) > CONTENT_PREVIEW_LENGTH else ''
        logger.info(f"已处理第{count + 1}条问题: {preview}{suffix}")

    @staticmethod
    def _generate_unique_id() -> str:
        """
        生成唯一ID

        Returns:
            str: UUID字符串
        """
        return str(uuid.uuid4())

    @staticmethod
    def _find_chunk_in_document(document: Document, chunk_id: Union[int, str]
    ) -> Optional[Any]:
        """
        在文档中查找指定的文本块

        Args:
            document: 文档对象
            chunk_id: 文本块ID

        Returns:
            Optional[Any]: 文本块对象或None
        """
        for chunk in document.chunks:
            if hasattr(chunk, 'id') and chunk.id == chunk_id:
                return chunk
        return None

    def generate_questions_for_document(
        self,
        document: Document,
        num_questions_per_chunk: int = DEFAULT_QUESTIONS_PER_CHUNK,
    ) -> List[Question]:
        """
        为文档生成问题

        Args:
            document: 目标文档对象
            num_questions_per_chunk: 每个文本块生成的问题数量 (默认3)

        Returns:
            List[Question]: 生成的问题列表
        """
        # 结果集合：保存生成的问题
        questions = []
        document_service = DocumentService(self.project_dir)

        for chunk_id in document.chunks:
            chunk = document_service.get_chunk(chunk_id)
            if not chunk:
                continue

            chunk_questions = self._generate_questions_for_chunk(
                document, chunk, num_questions_per_chunk
            )
            questions.extend(chunk_questions)

        # 将生成的问题持久化到磁盘
        self._save_questions(questions)
        self._update_document_status(document)

        return questions


    def generate_answers(
        self, questions: List[Question], with_chain_of_thought: bool
    ) -> List[Question]:
        """
        生成答案核心方法

        Args:
            questions: 要处理的问题列表
            with_chain_of_thought: 是否生成思维链

        Returns:
            List[Question]: 已更新状态的问题列表
        """
        updated_questions = []
        count = 1

        for question in questions:
            if question.answer:
                continue

            result = self._process_single_question(
                question, count, with_chain_of_thought
            )
            if result:
                updated_questions.append(result)
                count += 1

        return updated_questions

    def import_questions_from_xlsx(
            self,
            xlsx_file_path: str,
            document_id: Union[int, str],
            chunk_id: Union[int, str],
    ) -> List[Question]:
        """
        从XLSX导入问题

        Args:
            xlsx_file_path: XLSX文件路径
            document_id: 关联文档ID
            chunk_id: 关联文本块ID

        Returns:
            List[Question]: 导入的问题列表
        """
        imported_questions: List[Question] = []

        if not self._validate_xlsx_dependencies():
            raise ImportError("请安装 xlrd 库以导入 XLSX 文件")

        if not self._validate_xlsx_file(xlsx_file_path):
            return imported_questions

        workbook = self._load_xlsx_workbook(xlsx_file_path)
        if workbook is None:
            return imported_questions

        imported_questions = self._process_xlsx_rows(
            workbook, document_id, chunk_id, xlsx_file_path
        )

        logger.info(f"Excel导入完成: 共成功导入 {len(imported_questions)} 条问题")
        return imported_questions

    def get_questions_by_document_id(
            self, document_id: Union[int, str]
    ) -> List[Question]:
        """
        按文档ID获取问题

        Args:
            document_id: 文档ID

        Returns:
            List[Question]: 匹配的问题列表
        """
        questions = []

        if not os.path.exists(self.questions_dir):
            return questions

        for filename in os.listdir(self.questions_dir):
            if filename.endswith(".json"):
                question = self._load_question_from_file(filename, document_id)
                if question:
                    questions.append(question)

        return questions

    def _generate_questions_for_chunk(
        self, document: Document, chunk: Any, num_questions: int
    ) -> List[Question]:
        """
        为单个文本块生成问题

        Args:
            document: 文档对象
            chunk: 文本块对象
            num_questions: 生成的问题数量

        Returns:
            List[Question]: 生成的问题列表
        """
        # 调用LLM生成若干新问题
        generated_questions = self.llm_service.generate_questions(
            chunk.content,
            num_questions=num_questions,
            system_prompt=self.system_prompt,
        )

        questions = []
        for question_content in generated_questions:
            question = Question(
                document_id=document.id,
                chunk_id=chunk.id,
                content=question_content,
            )
            # 生成全局唯一ID用于存储
            question.id = self._generate_unique_id()
            questions.append(question)

        return questions

    def _save_questions(self, questions: List[Question]) -> None:
        """
        保存问题列表到磁盘

        Args:
            questions: 问题列表
        """
        for question in questions:
            question_path = os.path.join(
                self.questions_dir, f"{question.id}.json"
            )
            # 将问题对象写入磁盘
            save_file(question_path, question.to_dict(), base_dir=self.project_dir)

    def _update_document_status(self, document: Document) -> None:
        """
        更新文档状态为已生成问题

        Args:
            document: 文档对象
        """
        # 更新文档状态：已生成问题
        document.status = DOCUMENT_STATUS_QUESTIONS_GENERATED
        document_path = os.path.join(
            os.path.dirname(self.questions_dir),
            "documents",
            f"{document.id}.json",
        )
        save_file(document_path, document.to_dict(), base_dir=self.project_dir)

    def _process_single_question(
        self, question: Question, count: int, with_chain_of_thought: bool
    ) -> Optional[Question]:
        """
        处理单个问题的答案生成

        Args:
            question: 问题对象
            count: 当前处理计数
            with_chain_of_thought: 是否生成思维链

        Returns:
            Optional[Question]: 处理后的问题对象或None
        """
        # 使用文档服务定位文档与文本块信息
        document_service = DocumentService(self.project_dir)
        document = document_service.get_document(question.document_id)

        if not document:
            return None

        chunk = self._find_chunk_in_document(document, question.chunk_id)
        if not chunk:
            return None

        logger.info(f"正在处理问题 {count}: {question.content}")
        logger.debug("generating answer...")

        # 调用LLM生成答案
        question.answer = self.llm_service.generate_answer(
            question.content, chunk.content
        )

        if with_chain_of_thought:
            question.chain_of_thought = self.llm_service.generate_chain_of_thought(
                question.content, question.answer, chunk.content
            )

        question.status = QUESTION_STATUS_ANSWERED
        # 更新问题状态与持久化
        logger.debug("update status...")

        question_path = os.path.join(self.questions_dir, f"{question.id}.json")
        save_file(question_path, question.to_dict(), base_dir=self.project_dir)

        return question

    def _process_xlsx_rows(
        self,
        workbook: Any,
        document_id: Union[int, str],
        chunk_id: Union[int, str],
        xlsx_file_path: str,
    ) -> List[Question]:
        """
        处理XLSX工作表中的所有行

        Args:
            workbook: 工作簿对象
            document_id: 文档ID
            chunk_id: 文本块ID
            xlsx_file_path: 文件路径（用于错误信息）

        Returns:
            List[Question]: 导入的问题列表
        """
        imported_questions: List[Question] = []

        try:
            worksheet = workbook.active

            if worksheet.max_row == 0:
                logger.warning(f"警告: Excel文件为空 - {xlsx_file_path}")
                return imported_questions

            start_row = self._detect_data_start_row(worksheet)

            for row_idx, row in enumerate(
                worksheet.iter_rows(min_row=start_row, values_only=True),
                start=start_row,
            ):
                question = self._process_xlsx_row(
                    row, row_idx, document_id, chunk_id
                )
                if question:
                    imported_questions.append(question)

        except FileNotFoundError:
            logger.error(f"错误: 无法找到文件 - {xlsx_file_path}")
        except PermissionError:
            logger.error(f"错误: 没有权限访问文件 - {xlsx_file_path}")
        except Exception as e:
            logger.error(f"导入XLSX文件时出错: {str(e)}")
            logger.error(f"详细错误信息: {traceback.format_exc()}")

        return imported_questions

    def _process_xlsx_row(
        self,
        row: Tuple,
        row_idx: int,
        document_id: Union[int, str],
        chunk_id: Union[int, str],
    ) -> Optional[Question]:
        """
        处理单行XLSX数据

        Args:
            row: 行数据元组
            row_idx: 行索引
            document_id: 文档ID
            chunk_id: 文本块ID

        Returns:
            Optional[Question]: 创建的问题对象或None
        """
        if not row or all(cell is None for cell in row):
            return None

        if len(row) < 1 or row[0] is None:
            return None

        row_data = self._extract_row_data(row)
        if not row_data['content']:
            logger.warning(f"警告: 第{row_idx}行问题内容为空，跳过处理")
            return None

        question = self._create_question_from_row_data(
            row_data, document_id, chunk_id
        )

        self._print_import_progress(len([]), row_data['content'])

        return question

    def _create_question_from_row_data(
        self,
        row_data: dict,
        document_id: Union[int, str],
        chunk_id: Union[int, str],
    ) -> Question:
        """
        从行数据创建问题对象

        Args:
            row_data: 行数据字典
            document_id: 文档ID
            chunk_id: 文本块ID

        Returns:
            Question: 创建的问题对象
        """
        question = Question(
            document_id=document_id,
            chunk_id=chunk_id,
            content=row_data['content'],
            answer=row_data['answer'],
            chain_of_thought=row_data['chain_of_thought'],
        )

        question.status = self._determine_question_status(row_data)
        question.id = self._generate_unique_id()

        question_path = os.path.join(self.questions_dir, f"{question.id}.json")
        save_file(question_path, question.to_dict(), base_dir=self.project_dir)

        return question

    def _load_question_from_file(
        self, filename: str, document_id: Union[int, str]
    ) -> Optional[Question]:
        """
        从文件加载问题（如果匹配document_id）

        Args:
            filename: 文件名
            document_id: 目标文档ID

        Returns:
            Optional[Question]: 匹配的问题对象或None
        """
        file_path = os.path.join(self.questions_dir, filename)

        try:
            question_data = read_file(file_path, "json")
            question = Question.from_dict(question_data)  # type: ignore

            if question.document_id == document_id:
                return question
        except Exception as e:
            logger.warning(f"加载问题文件失败 {filename}: {e}")

        return None
